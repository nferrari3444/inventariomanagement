from typing import Any
from django.db.models.query import QuerySet
from django.contrib import messages
import pandas as pd
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from django_filters import rest_framework as filters
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import FormView
from django.db.models import Avg, Count, Exists, OuterRef
from django.db.models import Count, F, Value, Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from functools import reduce
from operator import or_
from django.utils.decorators import method_decorator
from django.views.generic import (CreateView, DeleteView, DetailView, ListView,
                                  UpdateView)
import json
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from datetime import datetime
from django.core import serializers
from django.views import generic
from annoying.functions import get_object_or_None
from django.http import HttpRequest, HttpResponseRedirect, JsonResponse, HttpResponse
from django.core.validators import validate_email
from django.core.exceptions import ValidationError, PermissionDenied
from .forms import SignUpForm , InboundForm, OutboundOrderForm, CotizationForm, OutboundDeliveryForm, InboundReceptionForm, TransferForm, TransferReceptionForm, CustomSetPasswordForm, CrudProductsForm
from .models import CustomUser,  StockMovements, DiffProducts, Product, WarehousesProduct, Tasks, Cotization
from .filters import StockFilterSet
from django.core.cache import cache
from django.utils.http import urlsafe_base64_encode
from django.core.mail import send_mail, BadHeaderError
from django.contrib.auth.views import PasswordResetView, PasswordContextMixin
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.contrib.messages.views import SuccessMessageMixin
import openpyxl
from django.db import transaction

# Create your views here.

class UserSignUpView(CreateView):
    model = CustomUser
    form_class = SignUpForm
    template_name = 'registration/signup_form.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'teacher'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('/')


@login_required
def home(request):
    return render(request, 'home.html')

def Register(request):

    if request.method == 'POST':
        username = request.POST.get('username','username')
        email = request.POST.get('email','email')

        try:
            validate_email(email)
        except ValidationError as e:
            messages.error(request, "Se debe agregar usuario y contraseña para Loguearse", extra_tags='login')


        password = request.POST.get('password', 'password')

        newUser = CustomUser.objects.create_user(username=username, email=email, password=password)

        newUser.save()

        return render(request, 'registration/login.html')
    
    else:
        return render(request, 'registration/register.html')
    
class PasswordResetConfirmView(PasswordContextMixin, FormView):
    form_class= CustomSetPasswordForm

def Login(request):
    
    page ='login'
    print('login view')
    if request.method == 'POST':
        email = request.POST.get('email','email')

        print('email is {}'.format(email))
        
        try:
            validate_email(email)
        except ValidationError as e:
            messages.error(request, 'El correo debe tener @', extra_tags='login')
            return render(request, 'registration/login.html')
        
        #username = CustomUser.objects.get(email=email)
        username = CustomUser.objects.filter(email=email).values_list('username', flat=True)
      #  print('username get is', username)
        print('username values_list is', username)

        print('request user is ', request.user)    
        password = request.POST.get('password','password')

        if len(username) == 0 | len(password) == 0:
            messages.error(request, "Se debe agregar usuario y contraseña para Loguearse", extra_tags='login')
            return render(request, 'registration/login.html')
        
        user = authenticate(request, email=  email , password=password)

        print('request user after auth is ', request.user)  
        print('user auth is ', user)  
        if user is not None:
            login(request, user)
            return redirect('/')
        
        else:
            messages.info(request, 'Usuario no existe. Se necesita dar de alta nuevo usuario')
            return render(request, 'registration/login.html')

    else:
        context = {'page' : page}
        return render(request, 'registration/login.html', context)
    

def Logout(request):
  
    logout(request)

    return redirect('/')


def getProducts(request):
 
    if cache.get('all') is not None:
        objects = cache.get('all')

        print('data in cache', objects)
    else:
        objects = Product.objects.all()
        categories = Product.objects.all().values('category').distinct()
        suppliers = Product.objects.all().values('supplier').distinct()
        warehouses = WarehousesProduct.objects.all().values('name').distinct()

        cache.set('products', objects)
        cache.set('categories', categories)
        cache.set('suppliers',suppliers) 
        cache.set('warehouses',warehouses)



    product = Product.objects.filter()
    product_data = Product.objects.values_list('name','product_id','internalCode')
  #  data = serializers.serialize('json', product_data, fields=('name','barcode', 'internalCode'))
    internalCodeJson = {product[2]: product[1] for product in product_data}

    productJson = {product[0]: product[1] for product in product_data}


    return JsonResponse({'products': productJson, 'codes':internalCodeJson}) #, 'warehouses': warehouses_json_data,'categories':  categories_json_data, 'supplier': suppliers_json_data})


def getProductsNames(request):
    
    print('autocomplete', request.GET.get('term'))
    if 'term' in request.GET:
        searchvalue = request.GET.get('term')
        if searchvalue.isdigit() == False:

            qs = Product.objects.filter(name__istartswith=searchvalue)
            titles = []
        
            for product in qs:
                titles.append(product.name)
         
            print(titles)       
            titles = list(set(titles)) 
            titles = sorted(titles)
            return JsonResponse(titles, safe=False)
        
        else:
            qs = Product.objects.filter(internalCode__istartswith= searchvalue)
            codes = []
        
            for product in qs:
                print('product in autocomplete is',product)
                print('internal code in autocomplete is', product.internalCode)
                codes.append(str(product.internalCode))
         
            print(codes)      
            codes = list(set(codes))  
            codes = sorted(codes)
            return JsonResponse(codes,safe=False)
            

def getProduct(request,productId):
    product = Product.objects.get(product_id=productId)
        
    print('product is', product)
    product_data = Product.objects.filter(product_id=productId).values_list('barcode','internalCode', 'name','category','supplier','quantity')
    
    print('product_data is', product_data)

    return JsonResponse({'product': list(product_data)})

def getProductWarehouse(request): #, productId, warehouse):

  
    productId = request.GET.get('productId',None)
    warehouse = request.GET.get('warehouse',None)
    quantity_input = request.GET.get('cantidad',None)

    cotizationCheck = request.GET.get('cotizationCheck',None)
    print('warehouse is', warehouse)
    print('productId is', productId)
    print('cotizationCheck', cotizationCheck)
    product = Product.objects.filter(product_id=productId)
    product_obj = Product.objects.get(product_id=productId)
    stockSecurity = product[0].stockSecurity

    if warehouse == None or warehouse == '' :
        response = JsonResponse({'error': "Seleccionar un deposito para continuar"})
        response.status_code = 403
        return response
    
      #product = Product.objects.filter(name=product_name)
    quantity = product[0].quantity
    if quantity == 0:
        response =  JsonResponse({'error': 'El producto {} no tiene Stock Disponible '.format(product[0].name)})
        response.status_code = 403
        return response

    
    if WarehousesProduct.objects.filter(product=product_obj, name=warehouse).exists() != True :
        print('da error en la vista')
        response =  JsonResponse({'error': 'El Producto ingresado no se encuentra en el Deposito {}'.format(warehouse)})
        response.status_code = 403
        return response

  

    if quantity_input:
        product_obj = Product.objects.get(product_id=productId)
        product_warehouse =  WarehousesProduct.objects.get(product=product_obj, name=warehouse)
        product_name = product_warehouse.product.name
        stock_warehouse_quantity = product_warehouse.quantity

        if int(stock_warehouse_quantity) < int(quantity_input):
            print('product_name is {}'.format(product_name))
            print('cantidad en deposito {}'.format(stock_warehouse_quantity))
            print('cantidad ingresada {}'.format(quantity_input))

            response =  JsonResponse({'error': 'La cantidad ingresada del producto {} es mayor a la cantidad en stock en el deposito {} '.format(product_name, warehouse)})
            response.status_code = 403
            return response
        
        if (quantity - int(quantity_input)) < stockSecurity:
            return JsonResponse({'message': 'La cantidad ingresada de {} deja al producto {} por debajo de su stock de Seguridad.'.format(quantity_input,product_name)})


    
        print('quantity input is', quantity_input)

    if product[0].hasOffer != None and cotizationCheck:
        product_name = product[0].name
        
        quantityOffer = product[0].quantityOffer
        

        if quantity - quantityOffer < stockSecurity * 1.1:
            product = Product.objects.filter(product_id=productId).values_list('barcode', 'internalCode', 'name',  'category','supplier','quantity', 'hasOffer' )
            #response =  JsonResponse({'error': 'El producto {} tiene cotización de Oferta y se encuentra cerca del Stock de Seguridad. El stock se encuentra reservado '.format(product_name)})
            #response =  JsonResponse({'success': 'El producto {} tiene cotización de Oferta y se encuentra cerca del Stock de Seguridad. El stock se encuentra reservado '.format(product_name)})

            return JsonResponse({'product': list(product), 'message': 'El producto {} tiene cotización de Oferta y queda por debajo del Stock de Seguridad.'.format(product_name)})
            #response.status_code = 403
            #return response
        else:
            print('el producto está en Oferta de cotizacion')
            print('product is', product)

            #product_warehouse = WarehousesProduct.objects.filter(product=product, name= warehouse).values_list('product__barcode','product__internalCode', 'product__name','name','location','product__category','product__supplier','quantity', 'product__hasOffer')

            product = Product.objects.filter(product_id=productId).values_list('barcode', 'internalCode', 'name',  'category','supplier','quantity', 'hasOffer' )

            return JsonResponse({'product': list(product), 'message': 'El producto {} tiene cotización de Oferta.'.format(product_name)})
            #response =  JsonResponse({)
            #response.status_code = 403
            #return response

    try:
        
        product = Product.objects.get(product_id=productId) #, warehouse= Warehouses.objects.get(name=warehouse))
        #product = Product.objects.get(name=product_name)

        product_warehouse_check = WarehousesProduct.objects.get(product=product, name=warehouse)
        
        #product = getProductWarehouse.objects.get(Product, name= warehouse)
        print('product is', product)
        # SAL ESCO (AXAL) x KG (BOLSA 25KG)
        #product_data = Product.objects.filter(product_id=productId, warehouse= Warehouses.objects.get(name=warehouse)).values_list('barcode','internalCode', 'name','warehouse','location','category','supplier','quantity', 'hasOffer')
        product_warehouse = WarehousesProduct.objects.filter(product=product, name= warehouse).values_list('product__barcode','product__internalCode', 'product__name','name','location','product__category','product__supplier','quantity', 'product__hasOffer')
        print('product_warehouse in view', product_warehouse)
        #print('product_data is', product_data)
        print('product_data is', product)

        return JsonResponse({'product': list(product_warehouse)})
    except WarehousesProduct.DoesNotExist:
        print('da error en la vista')
        response =  JsonResponse({'error': 'El Producto ingresado no se encuentra en el Deposito {}'.format(warehouse)})
        response.status_code = 403
        return response
    
    print('product has offer', product[0].hasOffer)
   

@login_required
def transferView(request):

    productNames = Product.objects.values_list('name', flat=True)
    print('extra_Field_count', request.POST.get('extra_field_count'))
    numberOfProducts = request.POST.get('extra_field_count')
    print('number of products in Inbound View is {}'.format(numberOfProducts))
    transferencia = StockMovements()
    # if this is a POST request we need to process the form data
    if request.method == "POST":
        # # create a form instance and populate it with data from the request:
        form = TransferForm(request.POST, user = request.user, extra = request.POST.get('extra_field_count'))
        # # check whether it's valid:
      #  print('form is valid', form.is_valid())
        print('request POST', request.POST)
       
        if form.is_valid():
      
            date  =   datetime.today().strftime('%Y-%m-%d')
        
            receptor = form.cleaned_data['receptor']
            warehouse_out = form.cleaned_data['warehouse']
            solicitante = request.user
            department = form.cleaned_data['department']
            motivoIngreso = 'Transferencia'
            # motivoIngreso = form.cleaned_data['motivoIngreso']
            actionType = 'Transferencia'
            observaciones = form.cleaned_data['observationsSolicitud']
       
            warehourse_inst = WarehousesProduct.objects.filter(name=warehouse_out).first()
            print('number of Products in form is {}'.format(numberOfProducts))
            datalist = []
            
            #products_list = [form.cleaned_data['product_{}'.format(i)] for i in range(0,int(numberOfProducts))]
            #print('products_list is {}'.format(products_list))
            #print('products_list exists:')
            
            print('Product that not exist in database are:')    
            #  warehouse= warehouse_out,
            task = Tasks.objects.create(date= date, receptor= receptor, warehouseProduct= warehourse_inst, issuer=solicitante,
                                        motivoIngreso=motivoIngreso,  actionType=actionType, department=department, observationsSolicitud = observaciones)
            
            for i in range(1,int(numberOfProducts) + 1):

                product = form.cleaned_data['product_{}'.format(i)]
                print('product in inbound view is {}'.format(product))
                internalCode = form.cleaned_data['internalCode_{}'.format(i)]
                
                # if Product.objects.filter(name__in=product, warehouse=warehouse_out).exists() != True:
                #     messages.error(request, 'El producto seleccionado {} no se encuentra en el deposito {}. Dar de alta el producto en el deposito para continuar'.format(product,warehouse_out), extra_tags='transfer')
                #     return redirect('/transfer/')
                    
                # Se modifica el objecto del producto. En lugar de buscar por nombre, se busca por internal Code
                productdb = Product.objects.get(internalCode = internalCode)
                #productdb = Product.objects.get(name= product)  #, warehouse= warehouse_out)
                warehouseProductdb = WarehousesProduct.objects.get(product=productdb, name=warehouse_out)
              
                quantity = form.cleaned_data['cantidad_{}'.format(i)]

                print('product is:', product)
                print('quantity is:', quantity)
                print('productId is:', productdb.product_id)
                #productdb = Product.objects.get(name= newproduct.product.name)

                # En la transferencia no se quita el stock del producto global
                #productToUpdate= Product.objects.filter(product_id= productdb.product_id) #  , warehouse=warehouse_out)
        
                #print('productToUpdate is:', productToUpdate)

                #productToUpdate.update(quantity = F('quantity') - quantity  )

               # depositProductUpdate = WarehousesProduct.objects.filter(product=productdb, name=warehouse_out)

               # depositProductUpdate.update(quantity = F('quantity') - quantity)

                newProduct = StockMovements(warehouseProduct =  warehouseProductdb, 
                            actionType = actionType,
                                        cantidad= quantity, task = task )
            

               
                WarehousesProduct.objects.create(product=productdb,
                                    name="En Transito", quantity=quantity, deltaQuantity = 0)

                datalist.append(newProduct)

           
            
          
            StockMovements.objects.bulk_create(datalist)     

            send_mail(
                subject='Transferencia de Productos entre Depositos',
                message= 'Transferencia de {} productos a depósito {}'.format(numberOfProducts,warehouse_out),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[receptor],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )
       
            return redirect('/tasks/')
            #return HttpResponseRedirect("/inbound/")

        else:
            print('request POST in invalid FOrm',request.POST)
            form = TransferForm(request.POST, user = request.user, extra = request.POST.get('extra_field_count'), invalidForm=True)
            return render(request, "transfer.html", {"form": form, "formInvalid" : True}) #
    # if a GET (or any other method) we'll create a blank form
    else:
        
        form = TransferForm(user=request.user)

    return render(request, "transfer.html", {"form": form, "formInvalid": False}) #, "products" :productNames})

@login_required
def transferEditTask(request,requested_id):
    pendingTask = get_object_or_None(Tasks, pk=requested_id)
    if pendingTask:
        products = pendingTask.stockmovements_set.all()

    tasks = Tasks.objects.filter(task_id=requested_id, status='Pending').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Products in OutboundDeliver view is', numberOfProducts)
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = TransferForm(request.POST,   user = request.user, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
            
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            
            # motivoEgreso = form.cleaned_data['motivoEgreso']
        
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()
            
           
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
            taskToUpdate.update(receptor=receptor, department=department, observationsSolicitud= observationsSolicitud,deliveryDate= deliveryDate)
            
            
            # task = Tasks.objects.get(task_id=requested_id)
            
            # print('products are:',products)
            # for i , product in enumerate(products):
            #     print('product in enumerate is:', product)
            #     form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name
            #     form.fields['cantidad_{}'.format(i)] = product.cantidad
            #     form.fields['internalCode_{}'.format(i)] = product.warehouseProduct.product.internalCode
            #     form.fields['barcode_{}'.format(i)] = product.warehouseProduct.product.barcode
                
                # internalCode = product.warehouseProduct.product.internalCode
                # quantity = form.fields['cantidad_{}'.format(i)]

                # #product_db = Product.objects.get(name= product.warehouseProduct.product.name)
                # product_db = Product.objects.get(internalCode= internalCode)
                # productWarehouse_db = WarehousesProduct.objects.get(product= product_db, name=warehouse)

                # productToUpdate= Product.objects.filter(product_id= product_db.product_id)
            
                # print('productToUpdate is:', productToUpdate)

                # productToUpdate.update(quantity = quantity) #, deltaQuantity = F('deltaQuantity') + diffQuantity  )
            
                # productWarehouseToUpdate = WarehousesProduct.objects.filter(product= product_db, name=warehouse)
                
                # productWarehouseToUpdate.update(quantity = quantity)

            print('form fields in view', form.fields)
            return redirect('/tasks/')
    else:
        print('pending task is:', pendingTask)
        form = TransferForm(instance= pendingTask , user = request.user, 
                                    initial={"task_id": requested_id })
    
    return render(request, "transferEdit.html", {"form": form, 'task_id': requested_id , "tasks": tasks})#, "numberOfProducts": len(productsToReceive)})

@login_required
def transferReceptionView(request, requested_id):

    #pendingTask = get_object_or_None(Tasks, pk=requested_id)
    pendingTask = Tasks.objects.filter(task_id=requested_id).first()
    print('request is:', request.method)
    print('request product data', request.POST.get('product_0'))
    product = request.POST.get('product')
    issuer = request.POST.get('issuer')
    #productsToReceive = pendingTask.stockmovements_set.all().values()
    # pendingTask:
    products = pendingTask.stockmovements_set.all()
    #task = Tasks.objects.filter(task_id = requested_id)
    tasks = Tasks.objects.filter(task_id=requested_id, actionType='Transferencia').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Produtos in Inbound Reception View is {}'.format(numberOfProducts))
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        
        form = TransferReceptionForm(request.POST, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
           #product = form.cleaned_data['product']
           # date  =   form.cleaned_data['date']
            
           # print(product)
           # print(date)
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            warehouseSalida = form.cleaned_data['warehouseSalida']
            print('warehouse in form is', warehouse)
            actionType = 'Confirma Transferencia'
            observations = form.cleaned_data['observationsConfirma']
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()





            datalist = []
            # task = Tasks.objects.create(date= date, receptor= receptor, warehouse= warehouse, issuer= issuer,
            #                             motivoIngreso=motivoIngreso,  actionType=actionType, department=department)
            
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
            taskToUpdate.update(status='Confirmed', actionType= actionType, observationsConfirma=observations, observationsSolicitud= observationsSolicitud, deliveryDate=deliveryDate)
            
            taskupdated = Tasks.objects.filter(task_id = requested_id).values_list('receptor', 'issuer','status', 'motivoIngreso','motivoEgreso','warehouseProduct__name','actionType')
            print('taskupdated in view is ', taskupdated)
            task = Tasks.objects.get(task_id=requested_id)
            form.fields['warehouse'] = warehouse
            #warehouseInTransit = Warehouses.objects.get(name='En Transito') 
            for i , product in enumerate(products):
                form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name                     #product.product.name
                form.fields['cantidad_{}'.format(i)] = product.cantidad
                netQuantity = form.cleaned_data['cantidadNeta_{}'.format(i)]
                
                quantity = form.fields['cantidad_{}'.format(i)]
                diffQuantity = int(quantity) - int(netQuantity)

                #productdb = Product.objects.filter(name= product.warehouseProduct.product.name)
                productdb = get_object_or_404(Product, product_id = product.warehouseProduct.product.product_id)
                print('productdb is', productdb)
                # productToUpdate= Product.objects.filter(product_id= newproduct.product_id, warehouse=warehouse)
                
                # Se debe actualizar el stock de producto en el depósito de salida, el stock en depósito En Transito, 
                # y el stock en el depósito de entrada.
                warehouseProductSalida = WarehousesProduct.objects.filter(product=productdb, name = warehouseSalida)
                warehouseProductSalida.update(quantity = F('quantity') - quantity) #, deltaQuantity = F('deltaQuantity') - diffQuantity)

                warehouseProductEnTransito = WarehousesProduct.objects.filter(product=productdb , name = 'En Transito')
                warehouseProductEnTransito.update(quantity = F('quantity') - quantity) #, deltaQuantity = F('deltaQuantity') - diffQuantity)


                # Se comenta productToUpdate porque ahora se va a actualizar el producto en el deposito
                #productToUpdate= Product.objects.filter(name= product.product.name, warehouse=warehouse)
                warehouseProduct = WarehousesProduct.objects.filter(product= productdb, name=warehouse)
                # productToUpdate = get_object_or_None(Product, name=product.product.name, warehouse=warehouse)
                print('productToUPdate is', warehouseProduct)
               

                # Se suman las cantidades que se recepcionan de los productos
                if warehouseProduct.exists():
                    warehouseProduct_data = WarehousesProduct.objects.get(product=productdb, name=warehouse)

                    #warehouseProductToUpdate = WarehousesProduct.objects.filter(product=productdb, name=warehouse)
                    warehouseProduct.update(quantity = F('quantity') + netQuantity, deltaQuantity = F('deltaQuantity') - diffQuantity)
                    
                    newProduct = StockMovements(warehouseProduct =  warehouseProduct_data  ,          #     product_data, 
                             actionType = actionType,
                                         cantidad= quantity, cantidadNeta=netQuantity, task = task )
                
                    datalist.append(newProduct)

                    if diffQuantity > 0 : #and motivoEgreso in ('Ventas','Planta de Armado'):
                        if DiffProducts.objects.filter(warehouseProduct= warehouseProduct_data).exists(): #, warehouse= warehouse).exists():
                        
                            DiffProducts.objects.filter(warehouseProduct= warehouseProduct_data).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity) #), warehouse=warehouse).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity)

                        else:
                            
                            DiffProducts.objects.create(warehouseProduct= warehouseProduct_data,  totalPurchase=quantity, totalQuantity= netQuantity, productDiff= diffQuantity)
                
                else:
                    
                    newProductInDeposit= WarehousesProduct.objects.create(product=productdb, name=warehouse, quantity = netQuantity, deltaQuantity=  diffQuantity)

                    newProductMovement = StockMovements(warehouseProduct = newProductInDeposit, 
                             actionType = actionType,
                                         cantidad= quantity, cantidadNeta=netQuantity, task = task )
                
                    datalist.append(newProductMovement)

                    if diffQuantity > 0 : #and motivoEgreso in ('Ventas','Planta de Armado'):
                        
                        if DiffProducts.objects.filter(warehouseProduct= newProductInDeposit).exists(): #, warehouse= warehouse).exists():
                        
                             DiffProducts.objects.filter(warehouseProduct= newProductInDeposit).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity) #), warehouse=warehouse).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity)

                        else:
                            DiffProducts.objects.create(warehouseProduct= newProductInDeposit,  totalPurchase=quantity, totalQuantity= netQuantity, productDiff= diffQuantity)
                    
                        
                    
                
                
                # productToDelete = Product.objects.filter(product_id= newproduct.product_id, warehouse='InTransit')
                #productToDelete = Product.objects.filter(name= product.product.name, warehouse= warehouseInTransit)
                productWarehouseToDelete = WarehousesProduct.objects.filter(product = productdb, name="En Transito")    
                print('productToUpdate is:', productWarehouseToDelete)

             
                productWarehouseToDelete.delete()

                # newProduct = StockMovements(product = productToUpdate, 
                #              actionType = actionType,
                #                          cantidad= quantity, cantidadNeta=netQuantity, task = task )
                
                # datalist.append(newProduct)
            
           # print('new_product is:', datalist)

            StockMovements.objects.bulk_create(datalist)     

            send_mail(
                subject='Confirmacion de Transferencia de Productos',
                message= 'Se confirma la transferencia de {} productos al depósito {}. La tarea fue confirmada por {}'.format(numberOfProducts,warehouse, receptor),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[issuer],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )

            return redirect('/tasks/')
            
    else:
       # form = OutboundOrderForm()
        form = TransferReceptionForm(instance= pendingTask , 
                                    initial={"task_id": requested_id })
        #form.task.queryset = Tasks.objects.filter(task_id = requested_id) 
        
    return render(request, "transferReception.html", {"form": form, 'task_id': requested_id , "tasks": tasks}) #, "numberOfProducts": len(productsToReceive)})

@login_required
def inboundView(request):

    print(request.user.role)
    productNames = Product.objects.values_list('name', flat=True)
    print('extra_Field_count', request.POST.get('extra_field_count'))
    numberOfProducts = request.POST.get('extra_field_count')
    print('number of products in Inbound View is {}'.format(numberOfProducts))
    nuevoIngreso = StockMovements()
    # if this is a POST request we need to process the form data
    if request.method == "POST":
        # # create a form instance and populate it with data from the request:
        form = InboundForm(request.POST, user = request.user, extra = request.POST.get('extra_field_count'))
        # # check whether it's valid:
        print('form is valid', form.is_valid())
       
        if form.is_valid():
        #     # process the data in form.cleaned_data as 
        #     product = form.cleaned_data['product']
            date  =   datetime.today().strftime('%Y-%m-%d')
           # print('form is:', form)

        #     print(product)
        #     print(date)
            
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            warehourse_inst = WarehousesProduct.objects.filter(name=warehouse).first()
            print('warehouse_inst in inboundView is ', warehourse_inst)

            solicitante =  request.user
            department = form.cleaned_data['department']
        #     cantidad = form.cleaned_data['cantidad']
        #     cantidadNeta = form.cleaned_data['cantidadNeta']
        #     deltaDiff =  cantidadNeta - cantidad
            motivoIngreso = form.cleaned_data['motivoIngreso']
            actionType = 'Nuevo Ingreso'
            observaciones = form.cleaned_data['observationsSolicitud']
          #  product = form.cleaned_data['product_23']
         #   print('product_23 is {}'.format(product))

            print('number of Products in form is {}'.format(numberOfProducts))
            datalist = []
            task = Tasks.objects.create(date= date, receptor= receptor, warehouseProduct= warehourse_inst, issuer=solicitante,
                                        motivoIngreso=motivoIngreso,  actionType=actionType, department=department, observationsSolicitud= observaciones)
            
            
            for i in range(1,int(numberOfProducts) + 1):

                product = form.cleaned_data['producto_{}'.format(i)]
                internalCode = form.cleaned_data['internalCode_{}'.format(i)]

                print('product in inbound view is {}'.format(product))
                print('internalCode in inbound view is {}'.format(internalCode))
                #product_ = product.strip()
                quantity = form.cleaned_data['cantidad_{}'.format(i)]
                productdb = Product.objects.get(internalCode= internalCode)
                warehouseProduct = WarehousesProduct.objects.filter(product= productdb, name=warehouse)

                if warehouseProduct.exists():                     #Product.objects.filter(name=product, warehouse=warehouse).exists():
                    print('el producto {} ya existe en el deposito {}'.format(product, warehouse))
                    # print('warehouse in line 430 is', warehouse)
                    # print('product in line 431 is', product)

                    warehouseproduct_db =  WarehousesProduct.objects.get(product=productdb, name=warehouse)          #Product.objects.get(name= product, warehouse=warehouse)
                    
                    # nuevoIngreso.barcode = form.cleaned_data['barcode_{}'.format(i)]
                    # nuevoIngreso.internalCode = form.cleaned_data['internalCode_{}'.format(i)]
                    

                    print('product is:', product)
                    print('quantity is:', quantity)

                    newWarehouseProduct = StockMovements(warehouseProduct = warehouseproduct_db, 
                             actionType = actionType,
                                         cantidad= quantity, task = task )
                
                    datalist.append(newWarehouseProduct)
                
                else:
                  
                    cantidad = form.cleaned_data['cantidad_{}'.format(i)]
                    
                    # Se comenta esto 31-10-2024 porque los productos no se mueven aun. Se pasa este codigo
                    # al momento que se confirma el producto. Finalmente se crea el producto con cantidad 0.

                    # productWarehouseNoStock = WarehousesProduct.objects.filter(product = productdb, name='No Stock')
                    # if productWarehouseNoStock.exists():                    
                    #     print('Se elimina el producto {} en el deposito No Stock'.format(productWarehouseNoStock))
                    #     productWarehouseNoStock.delete()
                    
                    # print('el producto {} no existe en el deposito {}'.format(product, warehouse))
                    # #newproduct_db = Product.objects.create(name=product, barcode=barcode,internalCode=internalCode,quantity=0, 
                    
                    # #                       warehouse= warehouse_obj, deltaQuantity=0, stockSecurity= stockSecurity, inTransit=True)

                    # #newproduct_db.save()
                    # 3-11-2024 Borramos el deposito del producto y le ingreso name=En Transito como
                    newWarehouseProduct_db = WarehousesProduct.objects.create(product=productdb, name=warehouse, quantity=0, deltaQuantity=0 , inTransit=True)
                    newWarehouseProduct_db.save()

                    newWarehouseProduct = StockMovements(warehouseProduct = newWarehouseProduct_db, actionType = actionType, cantidad= cantidad, task = task )
                    
                    datalist.append(newWarehouseProduct)
               
            print('new_product is:', datalist)

            StockMovements.objects.bulk_create(datalist)

            send_mail(
                subject='Nueva Solicitud de Ingreso de Materiales',
                message= 'Solicitud de Recepción de {} productos a depósito {} por motivo de {}'.format(numberOfProducts,warehouse,motivoIngreso),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[receptor],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )
        
        
            return redirect('/tasks/')
            #return HttpResponseRedirect("/inbound/")

    # if a GET (or any other method) we'll create a blank form
    else:
        
        form = InboundForm(user=request.user)

    return render(request, "inbound.html", {"form": form}) #, "products" :productNames})

@login_required
def inboundEditTask(request,requested_id):
    pendingTask = get_object_or_None(Tasks, pk=requested_id)
    if pendingTask:
        products = pendingTask.stockmovements_set.all()

    tasks = Tasks.objects.filter(task_id=requested_id, status='Pending').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Products in OutboundDeliver view is', numberOfProducts)
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = InboundForm(request.POST,   user = request.user, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
            
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            
            motivoIngreso = form.cleaned_data['motivoIngreso']
        
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()
            
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
            taskToUpdate.update(receptor=receptor, department=department, observationsSolicitud= observationsSolicitud, motivoIngreso=motivoIngreso, deliveryDate= deliveryDate)
            
            
            task = Tasks.objects.get(task_id=requested_id)
            
            # print('products are:',products)
            # for i , product in enumerate(products):
            #     print('product in enumerate is:', product)
            #     form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name
            #     form.fields['cantidad_{}'.format(i)] = product.cantidad
            #     form.fields['internalCode_{}'.format(i)] = product.warehouseProduct.product.internalCode
            #     form.fields['barcode_{}'.format(i)] = product.warehouseProduct.product.barcode
                
            #     internalCode = product.warehouseProduct.product.internalCode
            #     quantity = form.fields['cantidad_{}'.format(i)]

            #     product_db = Product.objects.get(internalCode= internalCode )

            #     productWarehouse_db = WarehousesProduct.objects.get(product= product_db, name=warehouse)

            #     productToUpdate= Product.objects.filter(product_id= product_db.product_id)
            
            #     print('productToUpdate is:', productToUpdate)

            #     productToUpdate.update(quantity = quantity) #, deltaQuantity = F('deltaQuantity') + diffQuantity  )
            
            #     productWarehouseToUpdate = WarehousesProduct.objects.filter(product= product_db, name=warehouse)
                
            #     productWarehouseToUpdate.update(quantity = quantity)

            # print('form fields in view', form.fields)
            return redirect('/tasks/')
    else:
        print('pending task is:', pendingTask)
        form = InboundForm(instance= pendingTask , user = request.user, 
                                    initial={"task_id": requested_id })
    
    return render(request, "inboundEdit.html", {"form": form, 'task_id': requested_id , "tasks": tasks})#, "numberOfProducts": len(productsToReceive)})

    
@login_required
def inboundReceptionView(request, requested_id):

    pendingTask = get_object_or_None(Tasks, pk=requested_id)
    print('request is:', request.method)
    print('request product data', request.POST.get('product_0'))
    product = request.POST.get('product')
    issuer = request.POST.get('issuer')
    #productsToReceive = pendingTask.stockmovements_set.all().values()
    if pendingTask:
        products = pendingTask.stockmovements_set.all()
    #task = Tasks.objects.filter(task_id = requested_id)
    tasks = Tasks.objects.filter(task_id=requested_id, actionType='Nuevo Ingreso').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Produtos in Inbound Reception View is {}'.format(numberOfProducts))
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        
        form = InboundReceptionForm(request.POST, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
      
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            actionType = 'Confirma Ingreso'
            motivoIngreso = form.cleaned_data['motivoIngreso']
            observations = form.cleaned_data['observationsConfirma']
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()
            
            # StockMovements.objects.create(product = product, date=date, department=department,
            #                             issuer=issuer, actionType = actionType, cantidad=cantidad,
            #                             motivoEgreso=motivoIngreso,status='Pending')
            
            datalist = []
            # task = Tasks.objects.create(date= date, receptor= receptor, warehouse= warehouse, issuer= issuer,
            #                             motivoIngreso=motivoIngreso,  actionType=actionType, department=department)
            
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
            taskToUpdate.update(status='Confirmed', observationsConfirma=observations, observationsSolicitud= observationsSolicitud, deliveryDate=deliveryDate)
            
            taskupdated = Tasks.objects.filter(task_id = requested_id).values_list('receptor', 'issuer','status', 'motivoIngreso','motivoEgreso','warehouseProduct__name','actionType')
            print('taskupdated in view is ', taskupdated)
            task = Tasks.objects.get(task_id=requested_id)
            
            for i , product in enumerate(products):
                form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name
                form.fields['cantidad_{}'.format(i)] = product.cantidad
                netQuantity = form.cleaned_data['cantidadNeta_{}'.format(i)]
                
                quantity = form.fields['cantidad_{}'.format(i)]
                diffQuantity = int(quantity) - int(netQuantity)

                newproduct = Product.objects.get(internalCode= product.warehouseProduct.product.internalCode) #, warehouse=warehouse)
                productWarehouse = WarehousesProduct.objects.get(product=newproduct, name=warehouse)
                
                # Este codigo anteriormente estaba en Inbound View. Si el producto se encuentra en deposito 
                # denominado No Stock => se elimina de ese depósito, y se da de alta como producto nuevo(create) 
                # en el deposito correspondiente.

                productWarehouseNoStock = WarehousesProduct.objects.filter(product = newproduct, name='No Stock')
                if productWarehouseNoStock.exists():                    
                    print('Se elimina el producto {} en el deposito No Stock'.format(productWarehouseNoStock))
                    productWarehouseNoStock.delete()
                    
                    print('el producto {} no existe en el deposito {}'.format(product, warehouse))
                    #newproduct_db = Product.objects.create(name=product, barcode=barcode,internalCode=internalCode,quantity=0, 
                    
                    #                       warehouse= warehouse_obj, deltaQuantity=0, stockSecurity= stockSecurity, inTransit=True)

                    #newproduct_db.save()
                    # newWarehouseProduct_db = WarehousesProduct.objects.create(product=newproduct, name=warehouse, quantity=quantity, deltaQuantity=0, inTransit=True)
                    # newWarehouseProduct_db.save()
               
               
                # Se actualizan las cantidades del proyecto
               

                print('El nuevo producto a recibir es {} en el deposito {}'.format(product.warehouseProduct.product.name, warehouse))
                productToUpdate= Product.objects.filter(product_id= newproduct.product_id) #, warehouse=warehouse)
        
                print('productToUpdate is:', productToUpdate)

                productToUpdate.update(quantity = F('quantity') + netQuantity) #, deltaQuantity = F('deltaQuantity') - diffQuantity , inTransit=False )
        

                productWarehouseToUpdate = WarehousesProduct.objects.filter(product= newproduct, name=warehouse)

                productWarehouseToUpdate.update(quantity = F('quantity') + netQuantity, deltaQuantity= F('deltaQuantity') - diffQuantity)

                # # Si el producto tiene inTransit=True, entonces el producto no estaba anteriormente en ningun deposito y se creo en InboundView en el 
                # # deposito correspondiente con la cantidad ingresada. 
                # if productWarehouseToUpdate[0].inTransit == True:
                #     productWarehouseToUpdate.update(quantity = netQuantity, deltaQuantity= F('deltaQuantity') - diffQuantity, inTransit=False)

                # else:
                #     productWarehouseToUpdate.update(quantity = F('quantity') + netQuantity, deltaQuantity= F('deltaQuantity') - diffQuantity)

                print('Se actualizó el producto {} en la tabla de productos así como en la tabal de warehouseProduct {} en el deposito {}'.format(productToUpdate, productWarehouseToUpdate,warehouse))
                newProduct = StockMovements(warehouseProduct = productWarehouse,         #newproduct, 
                             actionType = actionType,
                                         cantidad= quantity, cantidadNeta=netQuantity, task = task )
                

                if diffQuantity > 0: # and motivoIngreso in('Importación','Compra en Plaza'):
                    if DiffProducts.objects.filter(warehouseProduct=productWarehouse).exists(): #, warehouse= warehouse).exists():
                        
                        DiffProducts.objects.filter(warehouseProduct= productWarehouse).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity) #), warehouse=warehouse).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity)

                    else:
                        
                        DiffProducts.objects.create(warehouseProduct= productWarehouse,  totalPurchase=quantity, totalQuantity= netQuantity, productDiff= diffQuantity)
                
                datalist.append(newProduct)
            print('new_product is:', datalist)

            StockMovements.objects.bulk_create(datalist)     

            send_mail(
                subject='Confirmacion de Ingreso de Productos',
                message= 'Se confirma el Ingreso de {} productos al depósito {} por motivo de {}. La tarea fue confirmada por {}'.format(numberOfProducts,warehouse,motivoIngreso, receptor),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[issuer],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )
            return redirect('/tasks/')
            
    else:
       # form = OutboundOrderForm()
        form = InboundReceptionForm(instance= pendingTask ,
                                    initial={"task_id": requested_id })
        #form.task.queryset = Tasks.objects.filter(task_id = requested_id) 
        
    return render(request, "inboundReception.html", {"form": form, 'task_id': requested_id , "tasks": tasks}) #, "numberOfProducts": len(productsToReceive)})


@login_required
def inboundConfirmedView(request, requested_id):
   # pendingRequest = get_object_or_None(StockMovements, pk=requested_id)
    #print('pendingRequest product is:', pendingRequest.product)
    # confirmedTask = get_object_or_None(Tasks, pk=requested_id)
    confirmedTask = Tasks.objects.filter(task_id=requested_id, status__in = ('Confirmed','Cancelled'))
    print('confirmedTask is', confirmedTask)
    lastTask = Tasks.objects.filter(task_id= requested_id).first()
    actionType = lastTask.actionType
    taskStatus = lastTask.status
    print('lastTask is', lastTask)
    print('task actionType is', actionType)
    print('task status is', taskStatus)
    if taskStatus == 'Confirmed':
        title = 'Informacion Entrega de Productos Confirmada'
    elif taskStatus == 'Cancelled':
        title = 'Informacion Cancelación de Entrega de Productos'

    confirmedTask = Tasks.objects.filter(task_id=requested_id, status__in =('Confirmed', 'Cancelled'))

    firstTask = Tasks.objects.filter(task_id= requested_id).first()
    
    taskStatus = confirmedTask[0].status
    actionType = firstTask.actionType
    if taskStatus == 'Confirmed':
        firstMovement = firstTask.stockmovements_set.all().filter(actionType='Confirma Ingreso').first()
        actionType = firstMovement.actionType
        title = 'Informacion Ingreso de Productos Confirmada'
    elif taskStatus == 'Cancelled':
        firstMovement = firstTask.stockmovements_set.all().filter(actionType='Nuevo Ingreso').first()
        actionType = firstMovement.actionType
        title = 'Informacion Ingreso de Productos Cancelada'

    return render(request, 'inboundConfirmed.html', { 'task': confirmedTask[0] , 'actionType': actionType, 'title':title})

@login_required
def transferConfirmedView(request, requested_id):
   # pendingRequest = get_object_or_None(StockMovements, pk=requested_id)
    #print('pendingRequest product is:', pendingRequest.product)
    # confirmedTask = get_object_or_None(Tasks, pk=requested_id)
    
    confirmedTask = Tasks.objects.filter(task_id=requested_id, status__in =('Confirmed', 'Cancelled'))
    firstTask = Tasks.objects.filter(task_id= requested_id).first()
    
    # depositoSalida = firstMovement.warehouseProduct.name
    depositoEntrada = StockMovements.objects.filter(task__task_id=requested_id).last().warehouseProduct.name
   
    taskStatus = confirmedTask[0].status
    if taskStatus == 'Cancelled':
        firstMovement = firstTask.stockmovements_set.all().filter(actionType='Transferencia').first()
        actionType = firstMovement.actionType
        title = "Información de Transferencia de Productos Cancelada"
    elif taskStatus == 'Confirmed':
        firstMovement = firstTask.stockmovements_set.all().filter(actionType='Confirma Transferencia').first()
        actionType = firstMovement.actionType
        title = "Información de Transferencia de Productos Confirmada"
    return render(request, 'transferConfirmed.html', { 'task': confirmedTask[0] , 'actionType': actionType, 'warehouseIn': depositoEntrada, 'title':title})


@login_required
def outboundOrderView(request):
    print('request is:', request.method)
    print('request product data', request.POST.get('product'))
    product = request.POST.get('product')
    #issuer = request.POST.get('issuer')

    print('request user is ', request.user)
    context = {}
    numberOfProducts = request.POST.get('extra_field_count')

     # if this is a POST request we need to process the form data
    if request.method == "POST":
        
        form = OutboundOrderForm(request.POST, user = request.user, extra = request.POST.get('extra_field_count'))

        print('request POST FORM')
        print(request.POST)
        if form.is_valid():
            print('form is valid', form.is_valid())
            #print('form is', form)
            date  =   datetime.today().strftime('%Y-%m-%d')
           # print('form is:', form)

        #     print(product)
        #     print(date)
            
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            # solicitante = form.cleaned_data['issuer']
            warehourse_inst = WarehousesProduct.objects.filter(name=warehouse).first()
            department = form.cleaned_data['department']
            #     cantidad = form.cleaned_data['cantidad']
            #     cantidadNeta = form.cleaned_data['cantidadNeta']
            #     deltaDiff =  cantidadNeta - cantidad
            motivoEgreso = form.cleaned_data['motivoEgreso']
            observaciones = form.cleaned_data['observationsSolicitud']
            actionType = 'Nuevo Egreso'
            #form.fields['issuer'] = request.user
            solicitante = request.user
            #  product = form.cleaned_data['product_23']
            #   print('product_23 is {}'.format(product))

            print('number of Products in form is {}'.format(numberOfProducts))
            datalist = []
            task = Tasks.objects.create(date= date, receptor= receptor, warehouseProduct= warehourse_inst, issuer=solicitante,
                                        motivoEgreso=motivoEgreso,  actionType=actionType, department=department, observationsSolicitud= observaciones)
            
            print('form cleaned data', form.cleaned_data)
            for i in range(1,int(numberOfProducts) + 1 ):


                product = form.cleaned_data['producto_{}'.format(i)]
                internalCode = form.cleaned_data['internalCode_{}'.format(i)]

                product_db = Product.objects.get(internalCode= internalCode)
                productWarehouse_db = WarehousesProduct.objects.get(product= product_db, name=warehouse)
                # nuevoIngreso.barcode = form.cleaned_data['barcode_{}'.format(i)]
                # nuevoIngreso.internalCode = form.cleaned_data['internalCode_{}'.format(i)]
                quantity = form.cleaned_data['cantidad_{}'.format(i)]
               
                print('product is:', product)
                print('quantity is:', quantity)
                
                newProduct = StockMovements(warehouseProduct = productWarehouse_db, 
                             actionType = actionType,
                                         cantidad= quantity, task = task )

               
                datalist.append(newProduct)
                print('new_product is:', datalist)

            StockMovements.objects.bulk_create(datalist)

            send_mail(
                subject='Nueva Solicitud de Egreso de Materiales',
                message= 'Solicitud de Egreso por {} productos a depósito {} por motivo de {}. La solicitud fue ingresada por {}'.format(numberOfProducts,warehouse,motivoEgreso,solicitante),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[receptor],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )
        
            return redirect('/tasks/')
        else:

        #    context['form'] = form
            form = OutboundOrderForm(request.POST, user = request.user, extra = request.POST.get('extra_field_count'))

            return render(request, "outboundOrder.html", {"form": form})

    else:
        form = OutboundOrderForm(user = request.user)
        #context['form'] = OutboundOrderForm(user = request.user)
    
    return render(request, "outboundOrder.html", {"form": form})

class TaskListView(LoginRequiredMixin, generic.ListView):
    template_name = 'tasks.html'
    model = Tasks    

    
    def get_queryset(self) -> QuerySet[Any]:
        
        user_is_staff = self.request.user.is_staff
        if user_is_staff:
            self.tasks = Tasks.objects.all() # , actionType='Nueva Solicitud')

            return self.tasks

        
        print('user is superuser', self.request.user.is_superuser)
        user_group = self.request.user.groups.values_list('name',flat = True)[0].strip()

        print('user_group', user_group)
        if user_group == 'Tecnico':
            tasks = ["Transferencia","Confirma Transferencia"]
        elif user_group == 'Comercial':
            tasks = ["Nuevo Egreso", "Confirma Egreso"]
        else:
            tasks = ["Transferencia","Confirma Transferencia", "Nuevo Ingreso", "Confirma Ingreso", "Nuevo Egreso", "Confirma Egreso"]
        self.tasks = Tasks.objects.filter(actionType__in = tasks) #filter(status='Pending') # , actionType='Nueva Solicitud')

        return self.tasks
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get the context
        context = super(TaskListView, self).get_context_data(**kwargs)
        # Create any data and add it to the context
        context['tasks'] = self.tasks
        return context 

class FilteredListView(ListView):
    filterset_class = None

    def get_queryset(self):
        # Get the queryset however you usually would.  For example:
        queryset = super().get_queryset()
        queryset_stock = Product.objects.all()
        queryset_warehouse = WarehousesProduct.objects.all().order_by('product__internalCode')
        # Then use the query parameters and the queryset to
        # instantiate a filterset and save it as an attribute
        # on the view instance for later.
        print('request.GET in FilteredListView', self.request.GET)
        warehouse = self.request.GET.get('name',None)
        category = self.request.GET.get('category',None)
        supplier = self.request.GET.get('supplier',None)
        location = self.request.GET.get('location',None)
        
             
        if None not in (warehouse,category,supplier,location):
            self.filterset = self.filterset_class(self.request.GET, queryset=queryset_warehouse)
        else:
           self.filterset = self.filterset_class(self.request.GET, queryset=queryset_stock)

        #self.filterset = self.filterset_class(self.request.GET, queryset=queryset)
        # Return the filtered queryset
        return self.filterset.qs.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = super().get_queryset()
        # Pass the filterset to the template - it provides the form.
        queryset_stock = Product.objects.all()
        queryset_warehouse = WarehousesProduct.objects.all().order_by('product__internalCode')

        warehouse = self.request.GET.get('name',None)
        category = self.request.GET.get('category',None)
        supplier = self.request.GET.get('supplier',None)
        location = self.request.GET.get('location',None)

        print('warehouse isssss', warehouse)
      
        if None not in (warehouse,category,supplier, location):
            self.filterset = self.filterset_class(self.request.GET, queryset=queryset_warehouse)
        else:
           self.filterset = self.filterset_class(self.request.GET, queryset=queryset_stock)

        #self.filterset = self.filterset_class(self.request.GET, queryset=queryset)

        context['filterset'] = self.filterset
        return context


class StockListView(LoginRequiredMixin, FilteredListView, generic.ListView):

    queryset = Product.objects.all()
    filterset_class = StockFilterSet
    filter_backends = (filters.DjangoFilterBackend,)
    paginate_by = 20
    
    # paginate_by = 10
   # model = Product

    template_name = 'stock.html'
    
    
    def get_queryset(self):
        warehouse = self.request.GET.get('name', None)
        print('warehouse selected is', warehouse)
        supplier = self.request.GET.get('supplier',None)
        category = self.request.GET.get('category', None)
        location = self.request.GET.get('location', None)
        print('supplier selected is', supplier)

       # queryset = super().get_queryset()
       
        if category:
            category_filter = True
            categoryList = [category]
           # filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )

        else:
            categoryList = Product.objects.all().values('category').distinct().order_by('category')
           # filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )

        if location:
            locationList = [location]
        else:
            locationList = WarehousesProduct.objects.values_list('location',flat=True).distinct().order_by('location')

        if supplier:
            supplier_filter = True
            supplierList = [supplier]
           # filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )

        else:
            supplierList = Product.objects.all().values('supplier').distinct().order_by('supplier')
        #    filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )


        if warehouse:
            warehouseList = [warehouse]
         #   filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )

           
        else:
            warehouseList = WarehousesProduct.objects.values_list('name',flat=True).distinct().order_by('name')
        #    filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList )

        print('warehouse , category and supplier are {} {} {}'.format(warehouse,category,supplier))
        if None not in (warehouse,category,supplier, location) :
            filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, location__in= locationList, product__category__in= categoryList, product__supplier__in=supplierList).order_by('product__internalCode')

            
        else:
            filter = Product.objects.all()
            
        #filter = StockFilterSet(self.request.GET, queryset)
        return filter #.qs

    def get_context_data(self, **kwargs):
        
        warehouse = self.request.GET.get('name',None)
        print('warehouse selected get_context_data is', warehouse)
        supplier = self.request.GET.get('supplier')
        print('supplier selected get_context_data is', supplier)
        
        context = super().get_context_data(**kwargs)
        
        #queryset = self.get_queryset()
        #queryset = super().get_queryset()
        warehouse = self.request.GET.get('name', None)
        print('warehouse selected is', warehouse)
        supplier = self.request.GET.get('supplier',None)
        category = self.request.GET.get('category', None)
        location = self.request.GET.get('location', None)
        print('supplier selected is', supplier)

       # queryset = super().get_queryset()
       
        if category:
            category_filter = True
            categoryList = [category]
        else:
            categoryList = Product.objects.all().values('category').distinct().order_by('category')

        
        if location:
            locationList = [location]
        else:
            locationList = WarehousesProduct.objects.values_list('location',flat=True).distinct().order_by('location')


        if supplier:
            supplier_filter = True
            supplierList = [supplier]
        
        else:
            supplierList = Product.objects.all().values('supplier').distinct().order_by('supplier')

        if warehouse:
            warehouseList = [warehouse]
           
        else:
            warehouseList = WarehousesProduct.objects.values_list('name',flat=True).distinct().order_by('name')
          
           
        if None not in (warehouse,category,supplier, location) :
            filter = WarehousesProduct.objects.select_related('product').filter(name__in= warehouseList, product__category__in= categoryList, product__supplier__in=supplierList ).order_by('product__internalCode')

           
        else:
            filter = Product.objects.all()
           

        #filter = StockFilterSet(self.request.GET, queryset)
        context["filter"] = filter
        return context #filter #.qs
        
        
    
def filterProducts(request):
    
    
    warehouseSelection = False
    productSelection = False
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    page_num = request.GET.get('page')
    prods_codes = request.POST.getlist('arrcodes[]')
    prods_names = request.POST.getlist('arrnames[]')
    print('prods_codes is', prods_codes)
    print('prods_names is', prods_names)

    if is_ajax:
        
        data = dict()
       
        warehouse = request.GET.get('warehouse',None)
        category = request.GET.get('category',None)
        supplier = request.GET.get('supplier',None)
    
        #product = request.GET.get('product',None)
        codeSelection = request.POST.get('code')
        print('codeSelection is', codeSelection)

        checkbox= request.GET.get('checkbox',None)

        if checkbox: 
            productSelection = True              # Product.objects.select_related('warehouse').filter(inTransit=False)  
            filter_data = Product.objects.all()
            paginator = Paginator(filter_data, 20) # 6 employees per page

            page_num = request.GET.get('page')
            try:
                page_obj = paginator.page(page_num)
            except PageNotAnInteger:
                # if page is not an integer, deliver the first page
                page_obj = paginator.page(1)
            except EmptyPage:
                # if the page is out of range, deliver the last page
                page_obj = paginator.page(paginator.num_pages)

            # Se sustituye filter_data por page_obj
            # context = {'products' : filter_data}
            context = {'page_obj' : page_obj, 'product':productSelection}
    
            productList = Product.objects.all().values('name').distinct()
            categoryList = Product.objects.all().values('category').distinct()
            supplierList = Product.objects.all().values('supplier').distinct()
            warehouseList = WarehousesProduct.objects.all().values('name').distinct()
        
       # return render(request, 'stock.html',{'productList' : productList, 'supplierList':supplierList, 'warehouseList':warehouseList,   'categoryList':categoryList,  'page_obj': page_obj})
    
            data['html_table'] =  render_to_string('inject_table.html',
                                 context,
                                 request = request
                                 )
        
            #return render(request, 'stock.html', {'page_obj': page_obj})
            return JsonResponse(data)
        
    
        #print('product is',product)
        
        if prods_codes or prods_names: # product:
            #codeSelection = True if request.POST.get('code') == 'true' else False
            productSelection = True
            #if codeSelection == True:

            filter_data = Product.objects.filter(Q(internalCode__in=prods_codes) | Q(name__in= prods_names))
            print('filter_data', filter_data)

            context = {'page_obj' : filter_data, 'product': productSelection}

            data['html_table'] =  render_to_string('inject_table.html',
                                 context,
                                 request = request
                                 )
        
            #return render(request, 'stock.html', {'page_obj': page_obj})
            return JsonResponse(data)
                                                      
        
        if category and category != 'Total Categorias':
            category_filter = True
            categoryList = [category]
            productSelection = False
        else:
            categoryList = Product.objects.all().values('category').distinct()

        if supplier and supplier != 'Total Proveedores':
            supplier_filter = True
            supplierList = [supplier]
            productSelection = False
        
        else:
            supplierList = Product.objects.all().values('supplier').distinct()

    
        print('warehouse is', warehouse)
        print('category is', categoryList)
        print('supplier is', supplierList)

        if warehouse and warehouse != 'Total Depositos':
            # warehouse = WarehousesProduct.objects.get(name=warehouse)
            filter_data = WarehousesProduct.objects.select_related('product').filter(name=warehouse, product__category__in= categoryList, product__supplier__in=supplierList, inTransit=False )
            warehouseSelection = True
            productSelection = False
            # Product.objects.select_related('warehouse').filter(warehouse=warehouse, category__in =categoryList, supplier__in=supplierList, inTransit=False) 
        else:
            filter_data = WarehousesProduct.objects.filter(product__category__in= categoryList, product__supplier__in=supplierList, inTransit=False )
            warehouseSelection = True
            productSelection = False
            #filter_data = Product.objects.select_related('warehouse').filter(category__in =categoryList, supplier__in=supplierList, inTransit=False) 

    # data = serializers.serialize("json", Product.objects.filter(warehouse=warehouse, category=category, supplier=supplier).select_related('warehouse') )
        
        paginator = Paginator(filter_data, 20) # 6 employees per page

        page_num = request.GET.get('page')
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
            # if page is not an integer, deliver the first page
            page_obj = paginator.page(1)
        except EmptyPage:
            # if the page is out of range, deliver the last page
            page_obj = paginator.page(paginator.num_pages)

        # Se sustituye filter_data por page_obj
        # context = {'products' : filter_data}
        context = {'page_obj' : page_obj, 'warehouse':warehouseSelection, 'product': productSelection}
    
        productList = Product.objects.all().values('name').distinct()
        categoryList = Product.objects.all().values('category').distinct()
        supplierList = Product.objects.all().values('supplier').distinct()
        warehouseList = WarehousesProduct.objects.all().values('name').distinct()
        
       # return render(request, 'stock.html',{'productList' : productList, 'supplierList':supplierList, 'warehouseList':warehouseList,   'categoryList':categoryList,  'page_obj': page_obj})
    
        data['html_table'] =  render_to_string('inject_table.html',
                                 context,
                                 request = request
                                 )
        
        #return render(request, 'stock.html', {'page_obj': page_obj})
        return JsonResponse(data)

        #return render(request, 'stock.html', {'productList' : productList, 'supplierList':supplierList, 'warehouseList':warehouseList,   'categoryList':categoryList,   'page_obj': page_obj})
        
    
@login_required
def editDeliveryTask(request,requested_id):
    pendingTask = get_object_or_None(Tasks, pk=requested_id)
    if pendingTask:
        products = pendingTask.stockmovements_set.all()

    tasks = Tasks.objects.filter(task_id=requested_id, status='Pending').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Products in OutboundDeliver view is', numberOfProducts)
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = OutboundOrderForm(request.POST,   user = request.user, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
            
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            
            motivoEgreso = form.cleaned_data['motivoEgreso']
        
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()
            
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
            taskToUpdate.update(receptor=receptor, department=department, observationsSolicitud= observationsSolicitud, motivoEgreso=motivoEgreso, deliveryDate= deliveryDate)
            
            
            task = Tasks.objects.get(task_id=requested_id)
            
            # print('products are:',products)
            # for i , product in enumerate(products):
            #     print('product in enumerate is:', product)
            #     form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name
            #     form.fields['cantidad_{}'.format(i)] = product.cantidad
            #     form.fields['internalCode_{}'.format(i)] = product.warehouseProduct.product.internalCode
            #     form.fields['barcode_{}'.format(i)] = product.warehouseProduct.product.barcode
                
            #     internalCode = product.warehouseProduct.product.internalCode
            #     quantity = form.fields['cantidad_{}'.format(i)]

            #     product_db = Product.objects.get(internalCode= internalCode )

            #     productWarehouse_db = WarehousesProduct.objects.get(product= product_db, name=warehouse)

            #     productToUpdate= Product.objects.filter(product_id= product_db.product_id)
            
            #     print('productToUpdate is:', productToUpdate)

            #     productToUpdate.update(quantity = quantity) #, deltaQuantity = F('deltaQuantity') + diffQuantity  )
            
            #     productWarehouseToUpdate = WarehousesProduct.objects.filter(product= product_db, name=warehouse)
                
            #     productWarehouseToUpdate.update(quantity = quantity)

            # print('form fields in view', form.fields)
            return redirect('/tasks/')
    else:
        print('pending task is:', pendingTask)
        form = OutboundOrderForm(instance= pendingTask , user = request.user, 
                                    initial={"task_id": requested_id })
    
    return render(request, "outboundEdit.html", {"form": form, 'task_id': requested_id , "tasks": tasks})#, "numberOfProducts": len(productsToReceive)})


@login_required
def outboundDeliveryView(request, requested_id):
   # pendingRequest = get_object_or_None(StockMovements, pk=requested_id)
    #print('pendingRequest product is:', pendingRequest.product)
    pendingTask = get_object_or_None(Tasks, pk=requested_id)
    if pendingTask:
        products = pendingTask.stockmovements_set.all().filter(actionType='Nuevo Egreso')
    print('pendingTask is ', pendingTask)
    print('request is:', request.method)
    print('request product data', request.POST.get('producto_0'))
  
   # productsToReceive = pendingTask.stockmovements_set.all().values()
    #task = Tasks.objects.filter(task_id = requested_id)
    tasks = Tasks.objects.filter(task_id=requested_id, status='Pending').prefetch_related('stockmovements_set')

    numberOfProducts = request.POST.get('extra_field_count')
    print('number Of Products in OutboundDeliver view is', numberOfProducts)
     # if this is a POST request we need to process the form data
    if request.method == "POST":
        # create a form instance and populate it with data from the request:
        form = OutboundDeliveryForm(request.POST, instance= pendingTask, extra= request.POST.get('extra_field_count'))

        print('form is valid', form.is_valid())
      #  print('form is', form)
        print('form fields are', form.cleaned_data)

        # check whether it's valid:
        if form.is_valid():
            print('el form es valido')
           #product = form.cleaned_data['product']
           # date  =   form.cleaned_data['date']
            
           # print(product)
           # print(date)
            department = form.cleaned_data['department']
            issuer = form.cleaned_data['issuer']
            receptor = form.cleaned_data['receptor']
            warehouse = form.cleaned_data['warehouse']
            actionType = 'Confirma Egreso'
            motivoEgreso = form.cleaned_data['motivoEgreso']
            observations = form.cleaned_data['observationsConfirma']
            observationsSolicitud = form.cleaned_data['observationsSolicitud']
            deliveryDate = datetime.now().date()
            # StockMovements.objects.create(product = product, date=date, department=department,
            #                             issuer=issuer, actionType = actionType, cantidad=cantidad,
            #                             motivoEgreso=motivoIngreso,status='Pending')
            
            datalist = []
            # task = Tasks.objects.create(date= date, receptor= receptor, warehouse= warehouse, issuer= issuer,
            #                             motivoIngreso=motivoIngreso,  actionType=actionType, department=department)
            
            taskToUpdate = Tasks.objects.filter(task_id=requested_id)
        
            taskToUpdate.update(status='Confirmed', actionType=actionType, observationsConfirma=observations, observationsSolicitud= observationsSolicitud,  deliveryDate= deliveryDate)
            
            
            task = Tasks.objects.get(task_id=requested_id)
            
            for i , product in enumerate(products):
                form.fields['producto_{}'.format(i)] = product.warehouseProduct.product.name
                form.fields['cantidad_{}'.format(i)] = product.cantidad
                netQuantity = form.cleaned_data['cantidadNeta_{}'.format(i)]
                
                quantity = form.fields['cantidad_{}'.format(i)]
                diffQuantity = int(quantity) - int(netQuantity)

                product_db = Product.objects.get(name= product.warehouseProduct.product.name)

                productWarehouse_db = WarehousesProduct.objects.get(product= product_db, name=warehouse)

                productToUpdate= Product.objects.filter(product_id= product_db.product_id)
            
                print('productToUpdate is:', productToUpdate)

                productToUpdate.update(quantity = F('quantity') - netQuantity) #, deltaQuantity = F('deltaQuantity') + diffQuantity  )
            
                productWarehouseToUpdate = WarehousesProduct.objects.filter(product= product_db, name=warehouse)
                
                productWarehouseToUpdate.update(quantity = F('quantity') - netQuantity, deltaQuantity= F('deltaQuantity') + diffQuantity)

                newProduct = StockMovements(warehouseProduct = productWarehouse_db, 
                              actionType = actionType,
                                          cantidad= quantity, cantidadNeta=netQuantity, task = task )
                
                if diffQuantity > 0 : #and motivoEgreso in ('Ventas','Planta de Armado'):
                    if DiffProducts.objects.filter(warehouseProduct=productWarehouse_db).exists(): #, warehouse= warehouse).exists():
                        
                        DiffProducts.objects.filter(warehouseProduct= productWarehouse_db).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity) #), warehouse=warehouse).update(totalPurchase= F('totalPurchase') + quantity, totalQuantity= F('totalQuantity') + netQuantity, productDiff= F('productDiff') + diffQuantity)

                    else:
                        
                        DiffProducts.objects.create(warehouseProduct= productWarehouse_db,  totalPurchase=quantity, totalQuantity= netQuantity, productDiff= diffQuantity)
                
                datalist.append(newProduct)
            #print('new_product is:', datalist)
            print('form fields in view', form.fields)
            StockMovements.objects.bulk_create(datalist)     

            send_mail(
                subject='Confirmacion de Egreso de Materiales',
                message= 'Se confirma el Egreso de {} productos en el depósito {} por motivo de {}. La tarea fue confirmada por {}'.format(numberOfProducts,warehouse,motivoEgreso, receptor),
                from_email = settings.EMAIL_HOST_USER,
                recipient_list=[issuer],
                fail_silently=False,
                auth_user=None,
                auth_password=None,
                connection=None,
                html_message=None
            )

            return redirect('/tasks/')
           
    else:

        form = OutboundDeliveryForm(instance= pendingTask ,
                                    initial={"task_id": requested_id })
        #form.task.queryset = Tasks.objects.filter(task_id = requested_id) 
        
    return render(request, "outboundDelivery.html", {"form": form, 'task_id': requested_id , "tasks": tasks})#, "numberOfProducts": len(productsToReceive)})

@login_required
def outboundConfirmedView(request, requested_id):
   # pendingRequest = get_object_or_None(StockMovements, pk=requested_id)
    #print('pendingRequest product is:', pendingRequest.product)
    # confirmedTask = get_object_or_None(Tasks, pk=requested_id)
    confirmedTask = Tasks.objects.filter(task_id=requested_id, status__in = ('Confirmed','Cancelled'))
    print('confirmedTask is', confirmedTask)
    lastTask = Tasks.objects.filter(task_id= requested_id).first()
    actionType = lastTask.actionType
    taskStatus = lastTask.status
    print('lastTask is', lastTask)
    print('task actionType is', actionType)
    print('task status is', taskStatus)
    if taskStatus == 'Confirmed':
        title = 'Informacion Entrega de Productos Confirmada'
    elif taskStatus == 'Cancelled':
        title = 'Informacion Cancelación de Entrega de Productos'

    return render(request, 'outboundConfirmed.html', { 'task': confirmedTask[0], 'actionType':actionType, 'title': title})

def cancelTaskView(request, requested_id):

    # return render(request, 'cancelTaskModal.html')

    print('llega a cancelTaskView')
    print('requested_id is', requested_id)
    context = {'task': requested_id}
    print('request method is:', request.method)
    #if request.method == 'POST':
    taskToCancel = Tasks.objects.filter(task_id=requested_id) 
    taskToCancel.update(status='Cancelled')

    cancelledTask = get_object_or_None(Tasks, pk=requested_id)
    actionType = cancelledTask.actionType
    print('action Type for task id {} is {}'.format(requested_id, actionType))

    products = cancelledTask.stockmovements_set.all()

    for product in products:

        internalCode = product.warehouseProduct.product.internalCode
        productdb = Product.objects.get(internalCode = internalCode)
        if actionType == 'Transferencia':
           
            warehouse = 'En Transito'
            print('warehouse for producto to delete is', warehouse)
            print('producto con internal code a borrar', internalCode)
            

            newWarehouseProduct_db = WarehousesProduct.objects.get(product=productdb, name=warehouse)

        
            newWarehouseProduct_db.delete()
      

       

    

    return redirect('/tasks/')
    # else:
    #     return render(request, 'cancelTaskModal.html', context)

    

def finishTask(request, requested_id):


    print(requested_id)
    StockMovements.objects.filter(id=requested_id).update(status='Confirmed')
    #taskToConfirm.update(status='Confirmed')

    # taskToConfirm.save()
    
    return redirect('/tasks/')


class StockHistoryView(LoginRequiredMixin, generic.ListView):
    model = StockMovements

    template_name = 'stockhistory.html'

    def get_context_data(self, **kwargs):
        print('kwargs', self.kwargs)
        product_id = self.kwargs['product_id']
        product_name = Product.objects.filter(product_id=product_id)
        # Call the base implementation first to get the context
        context = super(StockHistoryView, self).get_context_data(**kwargs)
        # Create any data and add it to the context
        #movements = StockMovements.objects.filter(product=product_id)
        context['movements'] = StockMovements.objects.filter(warehouseProduct__product__name=
        product_name[0].name).select_related('task')
        
        context['product'] = product_name[0].name
        return context 


def export_excel(request, dimension):

    print('requestGET')
    print('dimension is',dimension)
   
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventario{}.xlsx"'.format(dimension)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'InventarioData{}'.format(dimension)

    if dimension == 'deposit':

        # Write header row
        header = ['Nombre', 'Codigo','Cantidad','Deposito','Categoria','Proveedor','Ubicacion','Stock de Seguridad']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        firstquery = WarehousesProduct.objects.all()
        queryset = firstquery.values_list('product__name', 'product__internalCode','quantity','name','product__category','product__supplier','location','product__stockSecurity').order_by('product__internalCode')
 
        #list(Product.objects.all().values('name', 'internalCode','quantity','warehouse__name','category','supplier','location','stockSecurity'))

        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = cell_value


    elif dimension == 'all':
         # Write header row
        header = ['Nombre', 'Codigo','CantidadTotal', 'Categoría','Proveedor','Stock de Seguridad']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        queryset = Product.objects.values_list('name', 'internalCode', 'quantity', 'category','supplier','stockSecurity').order_by('internalCode') #.annotate(CantidadTotal=Sum('quantity'))
    
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = cell_value

    workbook.save(response)

    
    return response

def handle_uploaded_file(file):

    file_data = pd.read_excel(file)
    file_data.columns = map(str.lower, file_data.columns)
    # Reemplazar los valores NaN por None
    file_data = file_data.astype(object).where(pd.notnull(file_data), None)
    return file_data
     
def cotizationDelete(request,cotization_id):
    
    Cotization.objects.filter(cotization_id=cotization_id).delete()
    products = Product.objects.filter(hasOffer=cotization_id)#.values_list('name', 'quantityOffer','priceOffer')    
   # products = get_object_or_404(Product, hasOffer=cotization_id)
    
    for product in products:
        product.hasOffer = None
        product.quantityOffer = None
        product.priceOffer = None

    

    #cotizations = Cotization.objects.all()

    return redirect('/new-cotization/')
    #return render(request, 'cotization.html', {'cotizations': cotizations})
    #return JsonResponse({'products': list(products)})


def cotizationView(request,cotization_id):

    products = Product.objects.filter(hasOffer=cotization_id)#.values_list('name', 'quantityOffer','priceOffer')    
   # products = get_object_or_404(Product, hasOffer=cotization_id)
    print('products is', products)
    context = {'products':products}
    return render(request, 'modal.html', context)
    #return JsonResponse({'products': list(products)})

def crudProducts(request,action):
    
    print('request.Files are')
    print(request.FILES)
    cotizations = Cotization.objects.all()
    title = action.capitalize() + ' Productos'
    print('action is', action)
    if request.method == 'POST':
        form = CrudProductsForm(request.POST, request.FILES)
        if form.is_valid():
            print('el form es valido')

            data = handle_uploaded_file(request.FILES['archivo'])
            
            register_date = datetime.now().date()
        
            numberOfProducts = len(data)
           

            if action == 'crear':
                products = []
                products_warehouse = []
                
                print('length of products to create is:', numberOfProducts)
                for i in range(0,len(data)):
                    try:
                        print('data product is:', data.iloc[i])
                        product_code = data.iloc[i][1] #['codigo interno']
                        product_quantity = data.iloc[i][3] # ['cantidad']
                        product_price = data.iloc[i][7] # ['precio']
                        deposit = data.iloc[i][9] # ['deposito']
                        product_barcode = data.iloc[i][2] #['codigo de barras']
                        name = data.iloc[i][0]# ['nombre']
                        category = data.iloc[i][4] #['categoria']
                        supplier = data.iloc[i][5] # ['proveedor']
                        stockSecurity = data.iloc[i][6] # ['stock seguridad']
                        currency = data.iloc[i][8] # ['moneda']
                        location = data.iloc[i][10] # ['ubicacion']

                        # El producto maestro se crea o actualiza
                        # obj, created = Product.objects.update_or_create(
                        # internalCode= product_code,
                        # defaults={"quantity": F('quantity') + product_quantity,
                        #         "barcode": product_barcode, 
                        #         "name": name, "category": category, 
                        #         "supplier": supplier, 
                        #         "stockSecurity": stockSecurity, "currency": currency})
                        
                        newProduct = Product.objects.filter(internalCode = product_code).exists()
                        print('newProduct is:', newProduct)
                        
                        if newProduct:
                            messages.error(request, 'El Producto con codigo {} ya existe en la base de datos'.format(product_code), extra_tags='product_exists')
            #                 messages.error(request, 'El Producto con codigo {} ya existe en la base de datos'.format(product_code), extra_tags='product_exists')
                        
                            return HttpResponseRedirect(reverse('productscrud', args=[action,]))    

                        
                        else:
                            print('product code is:', product_code)
                            print('product to add is:', newProduct)
                            
                            newProduct = Product.objects.create(name=name, internalCode= product_code, barcode= product_barcode, quantity= product_quantity,  price= product_price, category= category, supplier=supplier, stockSecurity=stockSecurity, currency=currency)
                            newProduct.save()
                            newProductInDeposit= WarehousesProduct(product= newProduct, name=deposit, quantity = product_quantity, location=location, deltaQuantity=0)
                            
                            # products.append(newProduct)
                            products_warehouse.append(newProductInDeposit)
                    
                    except ValidationError as e:
                        
                        messages.error(request, "Creacion de Producto con codigo {} es incorrecta. Chequear campos".format(product_code), extra_tags='product format')

                print('products to create object is:', products)
                WarehousesProduct.objects.bulk_create(products_warehouse)
               # Product.objects.bulk_create(products)
            
                messages.info(request, "Se crean {} productos".format(len(products_warehouse)))

            elif action == 'actualizar':
                for i in range(0,len(data)):
                    print('data product is:', data.iloc[i])
                    try:
                        product_code = data.iloc[i][0] #['codigo interno']
                        product_quantity = data.iloc[i][1] # ['cantidad']
                        product_price = data.iloc[i][2] # ['precio']
                        deposit = data.iloc[i][3] # ['deposito']


                        productToUpdate = Product.objects.filter(internalCode = product_code)
                        product_first = productToUpdate.first()
                        print('product first is', product_first)
                    
                        productToUpdate.update(quantity = F("quantity") + product_quantity,price =product_price)

                        productInDepositToUpdate = WarehousesProduct.objects.get(name=deposit, product__internalCode= product_code)
                        productInDepositToUpdate.quantity = product_quantity

                        productInDepositToUpdate.save()
                   
                    except WarehousesProduct.DoesNotExist:
                        productdb = None
                        messages.error(request, "Actualización de Productos con codigo {} es incorrecta. Chequear si el producto existe en deposito {}".format(product_code, deposit), extra_tags='product format')
                        return HttpResponseRedirect(reverse('productscrud', args=[action,]))    

                messages.info(request, "Se actualizan {} productos".format(len(data)))

            elif action == 'eliminar':
                for i in range(0,len(data)):
                    try:
                        product_code = data.iloc[i][0]
                        print('product_code is', product_code)
                        Product.objects.get(internalCode = product_code).delete()

                    except Product.DoesNotExist:
                        productdb = None
                        messages.error(request, "Eliminación de Producto con codigo {} es incorrecta. Chequear si el producto existe".format(product_code), extra_tags='product format')
                        return HttpResponseRedirect(reverse('productscrud', args=[action,]))    
                    
                messages.info(request, "Se eliminan {} productos".format(len(data)))
            
            elif action == "total":
                products_warehouse = []
                for i in range(0,len(data)):
                    print('data product is:', data.iloc[i])
                    try:
                        product_code = data.iloc[i][0] #['codigo interno']
                        print('product code is: ', product_code)
                        product_code_origin = data.iloc[i][1] # ['cantidad']
                        print('product code origin is: ', product_code_origin   )
                        category = data.iloc[i][3] # ['category']
                        supplier = data.iloc[i][4] # ['proveedor']
                        stock_raw = data.iloc[i][6]
                        stock = float(str(stock_raw).replace(',', '.')) if stock_raw not in [None, ''] else 0
                        

                        stockSecurity =  float(str(data.iloc[i][15]).replace(',', '.')) if data.iloc[i][15] not in [None, ''] else 0 # ['stock seguridad']
                        product_name = data.iloc[i][5] # ['nombre']
                        if (pd.isna(product_name) or product_name == '' or product_name is None) or pd.isna(category) or category == '' or category is None:
                            continue

                        warehouse_1 = 'Anaya 2710' # ['Anaya deposito']
                        ubication_warehouse_1 = data.iloc[i][8] # ['Anaya ubicacion']
                        quantity_warehouse_1_raw = data.iloc[i][7]
                        quantity_warehouse_1 = float(str(quantity_warehouse_1_raw).replace(',', '.')) if quantity_warehouse_1_raw not in [None, ''] else 0 # ['cantidad deposito 1
                        warehouse_2 = 'Crocker' # ['Crocker deposito']
                        ubication_warehouse_2 = data.iloc[i][10] # ['Crocker ubicacion']
                        quantity_warehouse_2_raw = data.iloc[i][9]
                        quantity_warehouse_2 = float(str(quantity_warehouse_2_raw).replace(',', '.')) if quantity_warehouse_2_raw not in [None, ''] else 0 # ['cantidad deposito 2
                        warehouse_3 = 'Joanico' # ['Juanico deposito']
                        ubication_warehouse_3 = data.iloc[i][12] # ['Juanico ubicacion']
                        quantity_warehouse_3_raw = data.iloc[i][11]
                        quantity_warehouse_3 = float(str(quantity_warehouse_3_raw).replace(',', '.')) if quantity_warehouse_3_raw not in [None, ''] else 0 # ['cantidad deposito 3
                        warehouse_4 = "In Transit" # ['In Transit deposito']
                        ubication_warehouse_4 = 'Transito' # ['ubicacion']
                        quantity_warehouse_4_raw = data.iloc[i][14]
                        quantity_warehouse_4 = float(str(quantity_warehouse_4_raw).replace(',', '.')) if quantity_warehouse_4_raw not in [None, ''] else 0 # ['cantidad deposito 4
                        
                        price_raw = data.iloc[i][28]
                        if price_raw is not None and price_raw != '':
                           
                            price_str = str(price_raw).strip('USD').strip().replace('-', '').replace('#¡REF! ', '').replace('#¡REF!', '')
                            # Handle European number format: remove dots (thousands separator) and replace comma with dot
                        
                            price = float(price_str.replace('.', '').replace(',', '.')) if price_str not in [None, ''] else None
                        else:
                            price = None  # ['precio de lista']

                        deposits = [warehouse_1, warehouse_2, warehouse_3, warehouse_4] 
                        locations = [ubication_warehouse_1, ubication_warehouse_2, ubication_warehouse_3, ubication_warehouse_4]
                        
                        quantities = [quantity_warehouse_1, quantity_warehouse_2, quantity_warehouse_3, quantity_warehouse_4]   
                        product_warehouse_quantities_list = list(zip(deposits, quantities,locations)) 
                          
                        newProduct = Product.objects.filter(internalCode = product_code).exists()
                        new_product_warehouse = WarehousesProduct.objects.filter(product= product_code, name__in= deposits).exists()
                        print('newProduct is:', newProduct)
                        print('newProductWarehouse is:', new_product_warehouse)
                        
                        if newProduct or new_product_warehouse:
                            productdb = Product.objects.get(internalCode = product_code)
                            productdb.category = category
                            productdb.supplier = supplier
                            productdb.stockSecurity = stockSecurity
                            productdb.quantity = stock
                            productdb.price = price
                            productdb.name = product_name
                            productdb.save()
                        #    newProduct.update(category= category, quantity = stock,  supplier=supplier, stock=stock, stockSecurity=stockSecurity, price=price, name= product_name)
                            
                            update_product_warehouse(deposits, productdb.internalCode, product_warehouse_quantities_list)
                           
    
            #                 messages.error(request, 'El Producto con codigo {} ya existe en la base de datos'.format(product_code), extra_tags='product_exists')
                          #  return HttpResponseRedirect(reverse('productscrud', args=[action,]))    

                        else:
                            newProduct = Product.objects.create(name=product_name, internalCode= product_code, barcode= product_code_origin, quantity= stock, price= price, category= category, supplier=supplier, stockSecurity=stockSecurity)
                            newProduct.save()

                            create_products_warehouse(deposits, newProduct.internalCode, product_warehouse_quantities_list)
                            #newProductInDeposit= WarehousesProduct(product= newProduct.internalCode, name=deposit, quantity = product_quantity, location=location, deltaQuantity=0)
                            
                            # products.append(newProduct)
                            #products_warehouse.append(newProductInDeposit)

                    except ValidationError as e:
                        
                        messages.error(request, "Creacion de Producto con codigo {} es incorrecta. Chequear campos".format(product_code), extra_tags='product format')

                messages.info(request, "Se crean o actualizan {} productos".format(len(data)))    
                return HttpResponseRedirect(reverse('productscrud', args=[action,]))    
                
    else:
        form = CrudProductsForm()
        
    
    return render(request, 'crudProducts.html', {'form': form, 'action': action,  'title': title})

def update_product_warehouse(deposits, product_code, product_warehouse_quantities):
    query_list = [Q(name__icontains=deposit) & Q(product__internalCode= product_code) for deposit in deposits]
    combined_query = reduce(or_, query_list)     
    results = WarehousesProduct.objects.filter(combined_query)  
    for i, product_warehouse in enumerate(results):
        if product_warehouse.name in product_warehouse_quantities[i][0]:
            product_warehouse.quantity = product_warehouse_quantities[i][1]
            product_warehouse.location = product_warehouse_quantities[i][2]
        product_warehouse.save()
    return results


def create_products_warehouse(deposits, product_code, product_warehouse_quantities):
    for i, deposit in enumerate(deposits):
        product_obj = Product.objects.get(internalCode= product_code)
        print('product warehouse quantities is:', product_warehouse_quantities[i]  )
        new_product_warehouse= WarehousesProduct(product= product_obj, name=product_warehouse_quantities[i][0], quantity = product_warehouse_quantities[i][1], location=product_warehouse_quantities[i][2], deltaQuantity=0)
        new_product_warehouse.save()
    



def newCotization(request):

    print('request.Files are')
    print(request.FILES)
    print('request object is', request)

    cotizations = Cotization.objects.all()
    if request.method == 'POST':
        form = CotizationForm(request.POST, request.FILES)
        if form.is_valid():
            print('el form es valido')
            data = handle_uploaded_file(request.FILES['archivo'])
            
            register_date = datetime.now().date()
            customer = form.cleaned_data['cliente']
            observations = form.cleaned_data['observaciones']
            numberOfProducts = len(data)

            cotization = Cotization.objects.create(date=register_date, customer = customer, numberOfProducts=numberOfProducts,observations=observations)

            for i in range(0,len(data)):
                product_code = data.iloc[i]['codigo']
                product_quantity = data.iloc[i]['cantidad']
                product_price = data.iloc[i]['precio']
                productdb = Product.objects.filter(internalCode = product_code).update(hasOffer=cotization, quantityOffer=product_quantity,priceOffer=product_price)

            messages.info(request, "Nueva Cotizacion de Oferta creada en forma exitosa", extra_tags="CotizacionOferta")
            #return HttpResponseRedirect('new-cotization/')
    else:
        form = CotizationForm()
    
    return render(request, 'cotization.html', {'form': form, 'cotizations':cotizations})